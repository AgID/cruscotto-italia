"""ETL RUNTS - Registro Unico Nazionale del Terzo Settore.

Fonte istituzionale: Ministero del Lavoro e delle Politiche Sociali.
Pubblicazione: D.Lgs 117/2017 art. 53 (pubblicita' legale RUNTS).
Licenza: CC BY 4.0 ex art. 52 c.2 D.Lgs 82/2005 (CAD) - open data di
default delle PA (in coerenza con le Linee Guida AgID Det. 183/2023).

URL ufficiale: https://servizi.lavoro.gov.it/runts/it-it/Lista-enti

Meccanismo di download (ASP.NET WebForms / DotNetNuke):
  1. GET sulla URL per ottenere VIEWSTATE / VIEWSTATEGENERATOR /
     EVENTVALIDATION + cookies di sessione.
  2. Identificazione DINAMICA del btnScaricaDoc associato alla riga della
     tabella gvEnti il cui titolo e' "Enti iscritti (formato Excel)".
     (Il numero ctlNN puo' cambiare se il Min. Lavoro aggiunge/rimuove righe.)
  3. POST form-urlencoded con __EVENTTARGET = nome del bottone.
  4. Risposta: XLSX binario ~12 MB con filename
     YYYYMMDD_iscritti_v1.0.xlsx (snapshot date nel filename).

Schema XLSX (verificato 15/05/2026):
  Foglio singolo 'Foglio1', 145.898 righe x 11 colonne (header bilingue IT/DE).

  Col 1: Codice fiscale  (string 11/16 char)
  Col 2: Repertorio       (int, id univoco RUNTS)
  Col 3: Denominazione    (string)
  Col 4: Sezione          (enum verbose, vedi SEZIONE_KEY)
  Col 5: Cognome+nome rapp. legale (string)
  Col 6: Rete             ('Si'/'No')
  Col 7: Comune Sede legale (string UPPERCASE) <- join key per ISTAT
  Col 8: Provincia Sede legale (sigla 2 char)
  Col 9: 5x1000           ('Si'/'No')
  Col 10: Data Iscrizione (gg/mm/aaaa)
  Col 11: vuota (da ignorare)

Schema output shard runts/<istat>.json:
{
  "_etl_version": "0.1.0",
  "_source": "RUNTS - Min. Lavoro e Politiche Sociali",
  "_source_url": "https://servizi.lavoro.gov.it/runts/it-it/Lista-enti",
  "_snapshot_date": "2026-05-15",
  "_generated_at": "ISO-8601",
  "kpi": {
    "n_totale": 446,
    "mix_sezione": {"APS": 198, "ODV": 124, "IS": 76, "ETS": 38, "EF": 6, "SMS": 4},
    "n_5x1000": 235,
    "pct_5x1000": 52.7,
    "n_rete_associativa": 1,
    "iscrizioni_per_anno": {"2022": 312, "2023": 74, "2024": 32, "2025": 19, "2026": 9}
  },
  "enti": [
    {"cf": "92123450753", "rep": "12345", "denom": "ASSOCIAZIONE XYZ APS",
     "sez": "APS", "rapp": "ROSSI MARIO", "rete": false, "x1000": true,
     "data_iscr": "2022-11-07"},
    ...   # cap 5000 ordinato data_iscr desc (Roma 6616, Milano 3716)
  ]
}

Comuni vuoti (n_totale=0): si genera comunque shard con enti=[] per
consistenza UX (la sezione non sara' 'null').

Pattern di esecuzione:
  1) download_xlsx() -> /tmp/cruscotto_runts/<YYYYMMDD>_iscritti_v1.0.xlsx
     skip se cache locale del giorno esiste e --skip-download
  2) parse_xlsx() -> list[dict] di tutti i 145k enti normalizzati
  3) load_canonical_lookups() -> nome_to_istat + canonical_istat
     (riusa la stessa logica di pnrr_progetti.py: bundle anagrafica +
     SPECIAL_NAMES per fusioni/bilingui)
  4) group_by_istat() -> dict[istat] = list[ente]
  5) build_shards() -> 7918 file runts/<istat>.json (cap 5000 enti)
  6) push_to_r2() -> list_objects_v2 + md5 diff + ThreadPool max_workers=24

Usage:
  python -m etl.sources.runts --target=r2
  python -m etl.sources.runts --target=local --outdir=dist/runts
  python -m etl.sources.runts --skip-download (riusa cache /tmp esistente)
"""

import argparse
import hashlib
import json
import os
import re
import sys
import time
import unicodedata
import urllib.parse
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

import requests
import structlog

from etl.lib import r2

log = structlog.get_logger(__name__)


# =========================================================================
# Costanti
# =========================================================================

RUNTS_URL = "https://servizi.lavoro.gov.it/runts/it-it/Lista-enti"
TARGET_ROW_TITLE = "Enti iscritti (formato Excel)"

CACHE_DIR = Path("/tmp/cruscotto_runts")
CACHE_DIR.mkdir(parents=True, exist_ok=True)

ETL_VERSION = "0.1.0"
SOURCE_LABEL = "RUNTS - Min. Lavoro e Politiche Sociali"

# Cap enti[] per shard per evitare payload >1MB su comuni metropolitani.
# Roma 6616, Milano 3716, Torino 2376: 5000 e' un compromesso.
# Ordinamento: data_iscr DESC (i piu' recenti per primi).
ENTI_CAP = 5000

# Mapping Sezione XLSX -> chiave breve per il JSON shard
SEZIONE_KEY = {
    "ORGANIZZAZIONI DI VOLONTARIATO": "ODV",
    "ASSOCIAZIONI DI PROMOZIONE SOCIALE": "APS",
    "ENTI FILANTROPICI": "EF",
    "IMPRESE SOCIALI": "IS",
    "SOCIETA' DI MUTUO SOCCORSO": "SMS",
    # Apostrofo tipografico (e' possibile che il file lo usi)
    "SOCIETA\u2019 DI MUTUO SOCCORSO": "SMS",
    "ALTRI ENTI DEL TERZO SETTORE": "ETS",
}

# Casi speciali RUNTS: nomi che il campo "Comune Sede legale" RUNTS scrive in
# forma differente dal bundle anagrafica. Si compone in 3 livelli:
#
# 1) Riuso mappe esistenti gia' verificate da altri ETL:
#      - veicoli.SPECIAL_ACI    (36 voci: bilingui + fusioni 2018-2024)
#      - pnrr_progetti.SPECIAL_NAMES (19 voci: bilingui FVG + Roma Capitale)
#    Pattern DRY: ogni fusione/bilingue censita una volta sola, riusabile.
#
# 2) Mapping RUNTS-specifici (73 voci aggiunte 2026-05-15 dopo il primo
#    log unmatched: 7.547 comuni distinct -> 7.620 atteso post-fix).
#    Codici verificati incrociando bundle anagrafica + storico ISTAT
#    Fusioni Comuni (https://www.istat.it/it/archivio/6789).
#
# 3) Bilingui Alto Adige (RUNTS non usa la barra: il bundle li ha come
#    "Bolzano/Bozen", il RUNTS scrive solo "BOLZANO").
#
# Da estendere dopo eventuali nuovi run osservando "runts_unmatched_comuni"
# nei log.
from etl.sources.veicoli import SPECIAL_ACI as _ACI_ALIASES
from etl.sources.pnrr_progetti import SPECIAL_NAMES as _PNRR_ALIASES

# Mapping aggiuntivi specifici RUNTS (non gia' in ACI o PNRR).
# Tutti i codici ISTAT sono stati VERIFICATI contro lookup/comuni-bundle.json
# il 2026-05-15 con lookup diretto per denominazione+provincia.
# Origine dei dati di verifica: il bundle stesso, fonte autoritativa.
_RUNTS_EXTRA_ALIASES: dict[str, str] = {
    # --- Bilingui Alto Adige (sono nel bundle come "X/Y", RUNTS scrive solo IT) ---
    "BOLZANO":                             "021008",  # Bolzano/Bozen
    "MERANO":                              "021051",  # Merano/Meran
    "BRESSANONE":                          "021011",  # Bressanone/Brixen
    "BRUNICO":                             "021013",  # Brunico/Bruneck
    "LAIVES":                              "021040",  # Laives/Leifers
    "VIPITENO":                            "021115",  # Vipiteno/Sterzing
    # --- Preposizioni mancanti / forme diverse nel RUNTS vs bundle ---
    "REGGIO CALABRIA":                     "080063",  # Reggio di Calabria
    "REGGIO EMILIA":                       "035033",  # Reggio nell'Emilia
    "CASSANO ALLO IONIO":                  "078029",  # Cassano all'Ionio (gia' in PNRR, ridondante OK)
    "CASTELNUOVO VAL DI CECINA":           "050011",  # Castelnuovo di Val di Cecina (PI)
    "CASTELNOVO NE'MONTI":                 "035016",  # Castelnovo ne' Monti (RE)
    "BARBERINO VAL D'ELSA":                "048054",  # -> Barberino Tavarnelle (FI, fus. 2019)
    # --- Bilingui FVG e fusioni FVG ---
    "TERZO DI AQUILEIA":                   "030120",  # Terzo d'Aquileia (UD)
    "FIUMICELLO":                          "030190",  # -> Fiumicello Villa Vicentina (UD, fus. 2018)
    "VALVASONE":                           "093053",  # -> Valvasone Arzene (PN, fus. 2015)
    "REANA DEL ROIALE":                    "030090",  # Reana del Rojale (UD)
    # --- Fusioni Trentino 2010-2018 ---
    "TAIO":                                "022230",  # -> Predaia (TN, fus. 2015)
    "TUENNO":                              "022230",  # -> Predaia (TN, fus. 2015)
    "CEMBRA":                              "022241",  # -> Cembra Lisignago (TN, fus. 2016)
    "ZAMBANA":                             "022167",  # -> San Michele all'Adige (TN, fus. 2018)
    "MOLINA DI LEDRO":                     "022229",  # -> Ledro (TN, fus. 2010)
    "SAN LORENZO IN BANALE":               "022231",  # -> San Lorenzo Dorsino (TN, fus. 2015)
    "SPERA":                               "022240",  # -> Castel Ivano (TN, fus. 2016)
    "PIEVEBOVIGLIANA":                     "043058",  # -> Valfornace (MC, fus. 2017)
    "PRIMIERO SAN MARTINO DI CASTRO":      "022245",  # Primiero S.M. di Castrozza (TN, RUNTS truncated)
    # --- Bolzano forme troncate dal RUNTS ---
    "CORTACCIA SULLA STRADA DEL VIN":      "021023",  # Cortaccia sulla Strada del Vino/Kurtatsch (BZ, RUNTS truncated)
    # --- Fusioni / rinomine Piemonte ---
    "MONTEMAGNO":                          "005077",  # Montemagno Monferrato (AT, gia' in ACI)
    "CASTELLINALDO":                       "004051",  # Castellinaldo d'Alba (CN)
    "CERESOLE D'ALBA":                     "004062",  # Ceresole Alba (CN, bundle senza apostrofo)
    "PECCO":                               "001318",  # -> Valchiusa (TO, fus. 2019)
    # --- Fusioni Veneto + Mantova ---
    "CRESPANO DEL GRAPPA":                 "026096",  # -> Pieve del Grappa (TV, fus. 2019)
    "VIRGILIO":                            "020071",  # -> Borgo Virgilio (MN, fus. 2014)
    "VILLA POMA":                          "020072",  # -> Borgo Mantovano (MN, fus. 2017)
    "FELONICA":                            "020061",  # -> Sermide e Felonica (MN, fus. 2017)
    "CARBONARA DI PO":                     "020072",  # -> Borgo Mantovano (MN, fus. 2017)
    "DRIZZONA":                            "019116",  # -> Piadena Drizzona (CR, fus. 2019)
    "PIADENA":                             "019116",  # -> Piadena Drizzona (CR, fus. 2019)
    "MIGLIARO":                            "038027",  # -> Fiscaglia (FE, fus. 2014)
    "MIGLIARINO":                          "038027",  # -> Fiscaglia (FE, fus. 2014)
    # --- Fusioni Lombardia ---
    "BREMBILLA":                           "016253",  # -> Val Brembilla (BG, fus. 2014)
    "VALLE MOSSO":                         "096088",  # -> Valdilana (BI, fus. 2019)
    "TRIVERO":                             "096088",  # -> Valdilana (BI, fus. 2019)
    "SOPRANA":                             "096088",  # -> Valdilana (BI, fus. 2019)
    "CAVALLASCA":                          "013206",  # -> San Fermo della Battaglia (CO, fus. 2017)
    "LENNO":                               "013252",  # -> Tremezzina (CO, fus. 2014)
    "MACCAGNO":                            "012142",  # -> Maccagno con Pino e Veddasca (VA, gia' in ACI)
    "SAN FEDELE INTELVI":                  "013254",  # -> Centro Valle Intelvi (CO, fus. 2017)
    "VERMEZZO":                            "015251",  # -> Vermezzo con Zelo (MI, fus. 2019)
    "RUINO":                               "018193",  # -> Colli Verdi (PV, fus. 2019)
    "MEL":                                 "025074",  # -> Borgo Valbelluna (BL, fus. 2019)
    # --- Fusioni Toscana ---
    "FIGLINE VALDARNO":                    "048052",  # -> Figline e Incisa Valdarno (FI, fus. 2014)
    "PERGINE VALDARNO":                    "051042",  # -> Laterina Pergine Valdarno (AR, fus. 2018)
    "LATERINA":                            "051042",  # -> Laterina Pergine Valdarno (AR, fus. 2018)
    "SCARPERIA":                           "048053",  # -> Scarperia e San Piero (FI, fus. 2014)
    "PIAN DI SCO":                         "051040",  # -> Castelfranco Piandiscò (AR, fus. 2014)
    "RIO MARINA":                          "049021",  # -> Rio (LI, fus. 2018)
    "VERGEMOLI":                           "046036",  # -> Fabbriche di Vergemoli (LU, fus. 2014)
    "GIUNCUGNANO":                         "046037",  # -> Sillano Giuncugnano (LU, fus. 2015)
    "SAN MARCELLO PISTOIESE":              "047024",  # -> San Marcello Piteglio (PT, fus. 2017)
    "CASCIANA TERME":                      "050040",  # -> Casciana Terme Lari (PI, fus. 2014)
    "CRESPINA":                            "050041",  # -> Crespina Lorenzana (PI, fus. 2014)
    # --- Fusioni Emilia-Romagna + Marche ---
    "SORBOLO":                             "034051",  # -> Sorbolo Mezzani (PR, fus. 2019)
    "ZIBELLO":                             "034050",  # -> Polesine Zibello (PR, fus. 2016)
    "MONTESCUDO":                          "099029",  # -> Montescudo-Monte Colombo (RN, fus. 2016)
    "SASSOCORVARO":                        "041071",  # -> Sassocorvaro Auditore (PU, fus. 2019)
    "MONTORO INFERIORE":                   "064121",  # -> Montoro (AV, fus. 2013; NB: provincia AV=064)
    # --- Veneto rinomine ---
    "COSTERMANO":                          "023030",  # Costermano sul Garda (VR, rinom. 2017)
    # --- Puglia/Campania/Calabria ---
    "CAPACCIO":                            "065025",  # Capaccio Paestum (SA, rinom. 2016)
    "SANNICANDRO GARGANICO":               "071049",  # San Nicandro Garganico (FG)
    "SANT'ANDREA APOSTOLO DELLO ION":      "079118",  # Sant'Andrea Apostolo dello Ionio (CZ, RUNTS truncated)
    "PRESICCE":                            "075098",  # -> Presicce-Acquarica (LE, fus. 2019)
    "CORIGLIANO CALABRO":                  "078157",  # -> Corigliano-Rossano (CS, fus. 2018)
    "PEDACE":                              "078156",  # -> Casali del Manco (CS, fus. 2017)
    # --- Sicilia / Sardegna ---
    "CALATAFIMI":                          "081003",  # Calatafimi-Segesta (TP)
    "RACCUIA":                             "083069",  # Raccuja (ME) -- RUNTS grafia con I, bundle con J
    "NIZZA SICILIA":                       "083061",  # Nizza di Sicilia (ME)
    "BARISARDO":                           "091005",  # Bari Sardo (NU)
    "IERZU":                               "091035",  # Jerzu (NU) -- RUNTS grafia con I, bundle con J
    # --- Lombardia Piacenza ---
    "PECORARA":                            "033049",  # -> Alta Val Tidone (PC, fus. 2018) -- NB: provincia PC=033, NON PV=018
    # --- VCO / Vicenza ---
    "FALMENTA":                            "103079",  # -> Valle Cannobina (VB, fus. 2019)
    "MOLVENA":                             "024126",  # -> Colceresa (VI, fus. 2019)
}

# SPECIAL_NAMES per RUNTS = union ACI + PNRR + RUNTS_EXTRA.
# Build deterministico: priorita' RUNTS_EXTRA > PNRR > ACI (in caso di chiavi
# duplicate, vince RUNTS_EXTRA che e' la mappa specifica della fonte).
SPECIAL_NAMES: dict[str, str] = {
    **_ACI_ALIASES,
    **_PNRR_ALIASES,
    **_RUNTS_EXTRA_ALIASES,
}


# =========================================================================
# Helpers
# =========================================================================

def _md5_file(p: Path) -> str:
    h = hashlib.md5()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def normalize(s: str) -> str:
    """Normalizza un nome comune per il match nome->ISTAT.

    Stessa logica di pnrr_progetti.normalize():
      1. Split su '/' (per nomi bilingui bundle tipo "Bolzano/Bozen")
      2. NFD-strip degli accenti
      3. Apostrofi (dritti e tipografici) rimossi
      4. Trattini sostituiti con spazio
      5. Uppercase + collapse di spazi multipli
    """
    if not s:
        return ""
    s = s.split("/")[0]
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode()
    s = s.replace("'", "").replace("\u2019", "")
    s = s.replace("-", " ")
    return " ".join(s.upper().split())


def _parse_date_ita(s: str) -> str | None:
    """Converte 'gg/mm/aaaa' in ISO 'aaaa-mm-gg'. None se non parsabile."""
    if not s or not isinstance(s, str):
        return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s.strip())
    if not m:
        return None
    g, mm, a = m.groups()
    try:
        d = datetime(int(a), int(mm), int(g))
        return d.strftime("%Y-%m-%d")
    except ValueError:
        return None


def _yesno(v) -> bool:
    """Converte 'Si'/'No' (e varianti) in bool. Default False."""
    if v is None:
        return False
    s = str(v).strip().upper()
    return s in ("SI", "SÌ", "S\u00CC", "YES", "TRUE", "1")


# =========================================================================
# FASE 1 — Download XLSX via PostBack ASP.NET
# =========================================================================

def download_xlsx(force: bool = False) -> Path:
    """Scarica l'XLSX RUNTS via ASP.NET PostBack. Ritorna path locale.

    Cache strategy: se esiste gia' un file YYYYMMDD_iscritti_v1.0.xlsx
    in CACHE_DIR con data odierna e force=False, riusa.
    """
    today = datetime.now().strftime("%Y%m%d")
    expected_path = CACHE_DIR / f"{today}_iscritti_v1.0.xlsx"
    if not force and expected_path.exists() and expected_path.stat().st_size > 1_000_000:
        log.info("runts_cache_hit",
                 path=str(expected_path),
                 size=expected_path.stat().st_size)
        return expected_path

    log.info("runts_download_start", url=RUNTS_URL)
    t0 = time.time()

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/125.0.0.0 Safari/537.36 CruscottoItalia/1.0",
        "Accept-Language": "it-IT,it;q=0.9,en;q=0.8",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    })

    # Step 1: GET per cookie + tokens
    r = session.get(RUNTS_URL, timeout=30)
    r.raise_for_status()
    html = r.text

    def _extract_token(name: str) -> str:
        m = re.search(
            rf'name="{name}"\s+id="{name}"\s+value="([^"]+)"',
            html,
        )
        if not m:
            raise RuntimeError(f"Token ASP.NET '{name}' non trovato in pagina")
        return m.group(1)

    vs = _extract_token("__VIEWSTATE")
    vsg = _extract_token("__VIEWSTATEGENERATOR")
    ev = _extract_token("__EVENTVALIDATION")

    # Step 2: identifica dinamicamente btnScaricaDoc per "Enti iscritti (formato Excel)"
    # Cerco la riga della tabella gvEnti che contiene il titolo target.
    # Il nome tabella e' dnn_ctrNNN_View_gvEnti dove NNN cambia tra deploy.
    table_match = re.search(
        r'<table[^>]*id="dnn_ctr\d+_View_gvEnti"[^>]*>(.+?)</table>',
        html, re.DOTALL,
    )
    if not table_match:
        raise RuntimeError("Tabella gvEnti non trovata nella pagina Lista-enti")
    table = table_match.group(1)

    rows = re.findall(r'<tr[^>]*>(.+?)</tr>', table, re.DOTALL)
    target_btn = None
    for row in rows:
        if TARGET_ROW_TITLE in row:
            m = re.search(
                r'name="(dnn\$ctr\d+\$View\$gvEnti\$ctl\d+\$btnScaricaDoc)"',
                row,
            )
            if m:
                target_btn = m.group(1)
                break
    if not target_btn:
        raise RuntimeError(
            f"Bottone download per '{TARGET_ROW_TITLE}' non trovato in gvEnti. "
            f"Il Min. Lavoro potrebbe aver modificato la struttura della pagina."
        )

    log.info("runts_target_button_found", button=target_btn)

    # Step 3: POST con __EVENTTARGET
    data = {
        "__EVENTTARGET": target_btn,
        "__EVENTARGUMENT": "",
        "__VIEWSTATE": vs,
        "__VIEWSTATEGENERATOR": vsg,
        "__EVENTVALIDATION": ev,
        "ScrollTop": "",
        "__dnnVariable": "",
    }
    r = session.post(
        RUNTS_URL,
        data=urllib.parse.urlencode(data),
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Referer": RUNTS_URL,
        },
        timeout=300,
    )
    r.raise_for_status()

    ct = r.headers.get("content-type", "")
    if "octet-stream" not in ct and "spreadsheet" not in ct:
        raise RuntimeError(
            f"Risposta non XLSX: content-type={ct} "
            f"(size={len(r.content)} bytes). Probabilmente la sessione "
            f"e' scaduta o il bottone e' cambiato."
        )

    cd = r.headers.get("content-disposition", "")
    fn_match = re.search(r'filename=([^;]+)', cd)
    filename = fn_match.group(1).strip('"') if fn_match else f"{today}_iscritti_v1.0.xlsx"

    # Snapshot date dal filename (formato YYYYMMDD_iscritti_v1.0.xlsx)
    if len(filename) >= 8 and filename[:8].isdigit():
        out_path = CACHE_DIR / filename
    else:
        out_path = expected_path

    out_path.write_bytes(r.content)
    elapsed = time.time() - t0
    log.info("runts_download_done",
             path=str(out_path),
             size=len(r.content),
             filename=filename,
             elapsed_s=round(elapsed, 1))
    return out_path


# =========================================================================
# FASE 2 — Parse XLSX
# =========================================================================

def parse_xlsx(xlsx_path: Path) -> list[dict]:
    """Parsa l'XLSX RUNTS in lista di dict. Ritorna ~145.898 record."""
    import openpyxl

    log.info("runts_parse_start", path=str(xlsx_path))
    t0 = time.time()

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active  # un solo foglio
    log.info("runts_xlsx_opened", sheet=ws.title, max_row=ws.max_row)

    records: list[dict] = []
    skipped_no_comune = 0
    skipped_no_sezione = 0
    header_seen = False

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Skip righe vuote
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        # Skip header (prima riga non-vuota: "Codice fiscale", ...)
        if not header_seen:
            first = str(row[0] or "").strip().lower()
            if "codice" in first and "fisc" in first:
                header_seen = True
                continue
            # se la prima riga non-vuota non e' l'header, comunque skip-defensive
            # solo se contiene parole tipiche header
            if any(isinstance(c, str) and c.strip().lower() in
                   ("repertorio", "denominazione", "sezione") for c in row):
                header_seen = True
                continue

        # Parse cella per cella
        cf = (str(row[0]).strip() if row[0] is not None else "")
        rep = (str(row[1]).strip() if row[1] is not None else "")
        denom = (str(row[2]).strip() if row[2] is not None else "")
        sez_raw = (str(row[3]).strip() if row[3] is not None else "")
        rapp = (str(row[4]).strip() if row[4] is not None else "")
        rete = _yesno(row[5])
        comune_raw = (str(row[6]).strip() if row[6] is not None else "")
        prov = (str(row[7]).strip() if row[7] is not None else "")
        x1000 = _yesno(row[8])
        data_iscr_raw = (str(row[9]).strip() if row[9] is not None
                         else "") if len(row) > 9 else ""

        if not comune_raw or comune_raw == "-":
            skipped_no_comune += 1
            continue

        sez_key = SEZIONE_KEY.get(sez_raw)
        if not sez_key:
            # Sezione "-" o ignota: skippiamo (handoff dice 12 righe nazionali)
            skipped_no_sezione += 1
            continue

        records.append({
            "cf": cf,
            "rep": rep,
            "denom": denom,
            "sez": sez_key,
            "rapp": rapp,
            "rete": rete,
            "comune": comune_raw,  # UPPERCASE come da XLSX
            "prov": prov,
            "x1000": x1000,
            "data_iscr": _parse_date_ita(data_iscr_raw),
        })

    wb.close()
    elapsed = time.time() - t0
    log.info("runts_parse_done",
             records=len(records),
             skipped_no_comune=skipped_no_comune,
             skipped_no_sezione=skipped_no_sezione,
             elapsed_s=round(elapsed, 1))
    return records


# =========================================================================
# FASE 3 — Load lookups bundle (nome -> ISTAT)
# =========================================================================

def load_lookups() -> tuple[dict[str, str], set[str]]:
    """Carica nome_to_istat + canonical_istat dal bundle anagrafica.

    Ritorna:
      - nome_to_istat: denominazione normalizzata -> istat_code
      - canonical_istat: set di tutti gli ISTAT validi (per build di shard
        anche per comuni con n_totale=0)
    """
    log.info("runts_lookups_loading")
    client = r2.get_r2_client()
    bucket = r2.get_bucket()
    obj = client.get_object(Bucket=bucket, Key="lookup/comuni-bundle.json")
    bundle = json.loads(obj["Body"].read())
    comuni = bundle.get("comuni", {})

    nome_to_istat: dict[str, str] = {}
    canonical_istat: set[str] = set()
    for istat, c in comuni.items():
        canonical_istat.add(istat)
        denom = c.get("denominazione", "")
        if denom:
            nome_to_istat[normalize(denom)] = istat

    # Casi speciali (bilingui senza "/" nel campo RUNTS, fusioni, etc.)
    # Validiamo che ogni ISTAT in SPECIAL_NAMES sia effettivamente nel bundle
    # canonico: se un codice e' sbagliato/inventato, lo escludiamo e logga
    # warning (vince la correttezza sulla copertura).
    invalid_aliases = []
    for n, istat in SPECIAL_NAMES.items():
        if istat not in canonical_istat:
            invalid_aliases.append((n, istat))
            continue
        nome_to_istat[normalize(n)] = istat
    if invalid_aliases:
        log.warning("runts_invalid_alias_istat",
                    n=len(invalid_aliases),
                    aliases=invalid_aliases[:10],
                    note="ISTAT non in bundle canonical: alias scartato. "
                         "Verifica i codici contro https://www.istat.it/"
                         "it/archivio/6789")

    log.info("runts_lookups_loaded",
             n_comuni=len(canonical_istat),
             nomi_normalizzati=len(nome_to_istat))
    return nome_to_istat, canonical_istat


# =========================================================================
# FASE 4 — Group by ISTAT + build KPI
# =========================================================================

def group_by_istat(records: list[dict],
                   nome_to_istat: dict[str, str]
                   ) -> tuple[dict[str, list[dict]], int]:
    """Raggruppa record per ISTAT. Ritorna (dict[istat] -> lista_enti, n_unmatched)."""
    grouped: dict[str, list[dict]] = defaultdict(list)
    unmatched: dict[str, int] = defaultdict(int)
    n_unmatched_records = 0

    for rec in records:
        key = normalize(rec["comune"])
        istat = nome_to_istat.get(key)
        if not istat:
            unmatched[rec["comune"]] += 1
            n_unmatched_records += 1
            continue
        grouped[istat].append(rec)

    if unmatched:
        # Log dei primi 20 comuni non matchati (per debug + futura estensione SPECIAL_NAMES)
        top_unmatched = sorted(unmatched.items(), key=lambda x: -x[1])[:20]
        log.warning("runts_unmatched_comuni",
                    n_distinct=len(unmatched),
                    n_records=n_unmatched_records,
                    top=top_unmatched)

    return dict(grouped), n_unmatched_records


def build_kpi(enti: list[dict]) -> dict:
    """Calcola KPI aggregati per un comune dato la lista enti."""
    n_totale = len(enti)
    if n_totale == 0:
        return {
            "n_totale": 0,
            "mix_sezione": {},
            "n_5x1000": 0,
            "pct_5x1000": 0.0,
            "n_rete_associativa": 0,
            "iscrizioni_per_anno": {},
        }

    mix_sezione: dict[str, int] = defaultdict(int)
    n_x1000 = 0
    n_rete = 0
    iscrizioni_per_anno: dict[str, int] = defaultdict(int)

    for e in enti:
        mix_sezione[e["sez"]] += 1
        if e["x1000"]:
            n_x1000 += 1
        if e["rete"]:
            n_rete += 1
        if e["data_iscr"]:
            anno = e["data_iscr"][:4]
            iscrizioni_per_anno[anno] += 1

    # Ordina mix_sezione per count desc
    mix_sorted = dict(sorted(mix_sezione.items(), key=lambda x: -x[1]))
    iscr_sorted = dict(sorted(iscrizioni_per_anno.items()))

    return {
        "n_totale": n_totale,
        "mix_sezione": mix_sorted,
        "n_5x1000": n_x1000,
        "pct_5x1000": round(100.0 * n_x1000 / n_totale, 1),
        "n_rete_associativa": n_rete,
        "iscrizioni_per_anno": iscr_sorted,
    }


def select_enti_for_shard(enti: list[dict], cap: int = ENTI_CAP) -> list[dict]:
    """Seleziona fino a `cap` enti, ordinati per data_iscr DESC (i piu' recenti).

    Per comuni piccoli ritorna tutti gli enti. Per Roma/Milano cappa a 5000.
    """
    # Sort: data_iscr desc (None alla fine come stringa vuota)
    enti_sorted = sorted(
        enti,
        key=lambda e: (e["data_iscr"] or ""),
        reverse=True,
    )
    return enti_sorted[:cap]


# =========================================================================
# FASE 5 — Build shards
# =========================================================================

def build_shard(istat: str,
                enti: list[dict],
                snapshot_date: str,
                generated_at: str) -> dict:
    """Costruisce il dict completo dello shard per un singolo comune."""
    kpi = build_kpi(enti)
    enti_out = select_enti_for_shard(enti)

    # Slim representation per il payload finale: rimuovi 'comune' e 'prov'
    # (gia' impliciti dal nome del file shard) per ridurre dimensione.
    enti_slim = [
        {
            "cf": e["cf"],
            "rep": e["rep"],
            "denom": e["denom"],
            "sez": e["sez"],
            "rapp": e["rapp"],
            "rete": e["rete"],
            "x1000": e["x1000"],
            "data_iscr": e["data_iscr"],
        }
        for e in enti_out
    ]

    shard = {
        "_etl_version": ETL_VERSION,
        "_source": SOURCE_LABEL,
        "_source_url": RUNTS_URL,
        "_snapshot_date": snapshot_date,
        "_generated_at": generated_at,
        "kpi": kpi,
        "enti": enti_slim,
    }

    # Nota: se il cap ha tagliato qualcosa, esplicitalo nel meta
    if len(enti) > ENTI_CAP:
        shard["_enti_truncated"] = True
        shard["_enti_total"] = len(enti)
        shard["_enti_cap"] = ENTI_CAP

    return shard


def build_all_shards(grouped: dict[str, list[dict]],
                     canonical_istat: set[str],
                     snapshot_date: str,
                     out_dir: Path) -> int:
    """Scrive uno shard JSON per ogni ISTAT canonico.

    Per comuni con 0 enti (~371), genera shard con kpi.n_totale=0 + enti=[]
    per consistenza UX (la sezione non sara' 'null' nel dashboard A1).
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    n_written = 0
    n_empty = 0
    n_capped = 0

    for istat in canonical_istat:
        enti = grouped.get(istat, [])
        if not enti:
            n_empty += 1
        if len(enti) > ENTI_CAP:
            n_capped += 1

        shard = build_shard(istat, enti, snapshot_date, generated_at)
        out_path = out_dir / f"{istat}.json"
        # JSON compatto (separator senza spazi) per ridurre size
        out_path.write_text(
            json.dumps(shard, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        n_written += 1

    log.info("runts_shards_written",
             total=n_written,
             empty=n_empty,
             capped=n_capped,
             outdir=str(out_dir))
    return n_written


# =========================================================================
# FASE 6 — Push R2 (pattern md5+ETag+ThreadPool)
# =========================================================================

def push_shards_to_r2(shard_dir: Path,
                      force_upload: bool = False) -> dict:
    """Push paralleli con skip via md5/ETag su prefix runts/."""
    if not shard_dir.exists():
        log.warning("runts_no_shard_dir_to_push", path=str(shard_dir))
        return {"uploaded": 0, "unchanged": 0, "errors": 0}

    import boto3 as _b3
    _client = _b3.client(
        "s3",
        endpoint_url=f"https://{os.environ['R2_ACCOUNT_ID']}.r2.cloudflarestorage.com",
        aws_access_key_id=os.environ["R2_ACCESS_KEY_ID"],
        aws_secret_access_key=os.environ["R2_SECRET_ACCESS_KEY"],
    )

    shard_files = sorted(shard_dir.glob("*.json"))

    remote_etag: dict[str, str] = {}
    try:
        _pag = _client.get_paginator("list_objects_v2")
        for _page in _pag.paginate(Bucket=r2.get_bucket(), Prefix="runts/"):
            for _o in _page.get("Contents", []):
                name = _o["Key"].split("/")[-1]
                etag = (_o.get("ETag") or "").strip('"').lower()
                remote_etag[name] = etag
        log.info("runts_shard_remote_listed", count=len(remote_etag))
    except Exception as e:
        log.warning("runts_shard_list_failed", error=str(e))

    to_upload: list[Path] = []
    if force_upload:
        to_upload = list(shard_files)
        log.info("runts_force_upload", count=len(to_upload))
    else:
        n_same = 0
        for sf in shard_files:
            rmd5 = remote_etag.get(sf.name)
            if rmd5 is None or _md5_file(sf) != rmd5:
                to_upload.append(sf)
            else:
                n_same += 1
        log.info("runts_md5_compared",
                 total=len(shard_files), unchanged=n_same,
                 to_upload=len(to_upload))

    def _upload_one(sf: Path) -> str:
        r2.upload_file(sf, f"runts/{sf.name}",
                       content_type="application/json")
        return sf.name

    uploaded = 0
    errors = 0
    with ThreadPoolExecutor(max_workers=24) as ex:
        futures = {ex.submit(_upload_one, sf): sf for sf in to_upload}
        for f in as_completed(futures):
            try:
                f.result()
                uploaded += 1
                if uploaded % 200 == 0:
                    log.info("runts_push_progress",
                             uploaded=uploaded, total=len(to_upload))
            except Exception as e:
                errors += 1
                log.error("runts_upload_failed", error=str(e))

    log.info("runts_push_done",
             uploaded=uploaded,
             unchanged=len(shard_files) - len(to_upload),
             errors=errors)
    return {
        "uploaded": uploaded,
        "unchanged": len(shard_files) - len(to_upload),
        "errors": errors,
    }


# =========================================================================
# Main CLI
# =========================================================================

def main() -> int:
    ap = argparse.ArgumentParser(
        description="ETL RUNTS (Registro Unico Nazionale Terzo Settore)",
    )
    ap.add_argument("--target", choices=["local", "r2"], default="local",
                    help="local: scrive solo in --outdir; r2: scrive shard + push R2")
    ap.add_argument("--outdir", default="/var/www/cruscotto-italia/data/runts",
                    help="Directory shard locali (default: /var/www/cruscotto-italia/data/runts)")
    ap.add_argument("--skip-download", action="store_true",
                    help="Riusa cache locale /tmp/cruscotto_runts/ (non scarica)")
    ap.add_argument("--force-download", action="store_true",
                    help="Re-download anche se cache odierna presente")
    ap.add_argument("--force-upload", action="store_true",
                    help="Push R2 di tutti gli shard senza md5 check")
    ap.add_argument("--xlsx-path", default="",
                    help="Path esplicito a un XLSX gia' scaricato (override cache)")
    args = ap.parse_args()

    log.info("runts_etl_start", target=args.target)
    t_start = time.time()

    # FASE 1: download (o riuso cache)
    if args.xlsx_path:
        xlsx_path = Path(args.xlsx_path)
        if not xlsx_path.exists():
            log.error("runts_xlsx_not_found", path=str(xlsx_path))
            return 1
        log.info("runts_xlsx_explicit", path=str(xlsx_path))
    elif args.skip_download:
        # Cerca l'XLSX piu' recente in CACHE_DIR
        candidates = sorted(CACHE_DIR.glob("*_iscritti_v*.xlsx"), reverse=True)
        if not candidates:
            log.error("runts_cache_empty",
                      cache_dir=str(CACHE_DIR),
                      note="usa --xlsx-path o rimuovi --skip-download")
            return 1
        xlsx_path = candidates[0]
        log.info("runts_cache_reused", path=str(xlsx_path))
    else:
        xlsx_path = download_xlsx(force=args.force_download)

    # Snapshot date dal nome file (YYYYMMDD_iscritti_v1.0.xlsx)
    fn = xlsx_path.name
    if len(fn) >= 8 and fn[:8].isdigit():
        snapshot_date = f"{fn[:4]}-{fn[4:6]}-{fn[6:8]}"
    else:
        snapshot_date = datetime.now().strftime("%Y-%m-%d")
    log.info("runts_snapshot_date", date=snapshot_date)

    # FASE 2: parse XLSX
    records = parse_xlsx(xlsx_path)
    if not records:
        log.error("runts_no_records_parsed")
        return 1

    # FASE 3: load lookups
    nome_to_istat, canonical_istat = load_lookups()

    # FASE 4: group by ISTAT
    grouped, n_unmatched = group_by_istat(records, nome_to_istat)
    pct_unmatched = round(100.0 * n_unmatched / len(records), 3)
    log.info("runts_grouped",
             records=len(records),
             comuni_distinti=len(grouped),
             unmatched=n_unmatched,
             pct_unmatched=pct_unmatched)
    if pct_unmatched > 1.0:
        log.warning("runts_unmatched_high",
                    pct=pct_unmatched,
                    note="estendi SPECIAL_NAMES con i comuni in top_unmatched")

    # FASE 5: build shards
    out_dir = Path(args.outdir)
    n_written = build_all_shards(grouped, canonical_istat,
                                  snapshot_date, out_dir)

    # FASE 6: push R2 (solo se target=r2)
    if args.target == "r2":
        result = push_shards_to_r2(out_dir, force_upload=args.force_upload)
        log.info("runts_r2_push_result", **result)

    elapsed = time.time() - t_start
    log.info("runts_etl_done",
             elapsed_s=round(elapsed, 1),
             shards=n_written,
             snapshot=snapshot_date)
    return 0


if __name__ == "__main__":
    sys.exit(main())
