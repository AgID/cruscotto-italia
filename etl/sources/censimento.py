"""ETL ISTAT - Basi Territoriali + Variabili censuarie 2021.

Fonte: ISTAT - Censimento permanente popolazione 2021.
Licenza: CC BY 3.0 IT (standard ISTAT).
URL pagina: https://www.istat.it/notizia/basi-territoriali-e-variabili-censuarie/

Composizione fonte:
- 20 ZIP regionali shapefile sezioni di censimento 2021 (WGS84 UTM Zona 32N):
    https://www.istat.it/storage/cartografia/basi_territoriali/2021/R<NN>_21.zip
- 1 ZIP nazionale variabili censuarie sezioni 2021 (XLSX per regione):
    https://esploradati.istat.it/databrowser/DWL/PERMPOP/SUBCOM/Dati_regionali_2021.zip
- 1 ZIP aree subcomunali 2021 (solo ~43 capoluoghi, layer overlay opzionale):
    https://www.istat.it/wp-content/uploads/2025/04/ASC_21.zip

Edizione: dati definitivi pubblicati 14/05/2026 (pagina ISTAT aggiornata).

Copertura: 7896/7896 comuni (TN/BZ inclusi, a differenza del Catasto AdE).
Strategia: ETL local-first, output flat in data/censimento_full/<istat>.geojson
(geometrie + 122 variabili per sezione) + un solo file di aggregati
data/censimento/aggregati.json (dict {istat: kpi_comune+distribuzioni})
letto da dashboard.py per la sezione "censimento" nel comune A1.

Schema 122 variabili (dal file TRACCIATO_2021 ufficiale ISTAT):
  - P1-P3: popolazione totale + sesso
  - P14-P29: popolazione totale per fascia eta 5 anni
  - P30-P45: popolazione maschi per fascia eta 5 anni
  - P67-P82: popolazione femmine per fascia eta 5 anni
  - P83-P100: titolo di studio (totale, maschi, femmine x 5 livelli)
  - P101-P103: occupati 15-64 (totale, maschi, femmine)
  - IT1-IT12: italiani per fascia eta (0-14, 15-64, 65+) x sesso x occupati
  - ST1-ST33: stranieri (totali, UE/extra-UE, fasce eta, sesso, occupati)
  - PF1, PF3-PF8: famiglie per numero componenti
  - A2, A3, A8: abitazioni (occupate, vuote, totali)
  - E3: edifici residenziali

Output:
- data/censimento_full/<istat>.geojson : 7896 file, FeatureCollection con
  geometrie (Polygon EPSG:4326) + properties.vars (dict 122 variabili)
- data/censimento/aggregati.json : 1 file con dict aggregati comune-level
  per il dashboard A1 (kpi_comune + distribuzioni)

Aggiornamento: annuale, allineato al rilascio ISTAT (tipicamente aprile).
"""
from __future__ import annotations

import argparse
import io
import json
import os
import sys
import zipfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import structlog

from etl.lib import manifest

log = structlog.get_logger()

# ═════════════════════════════════════════════════════════════════════════
# Costanti fonte
# ═════════════════════════════════════════════════════════════════════════

SOURCE_LABEL = "ISTAT - Basi Territoriali + Variabili censuarie 2021"
SOURCE_PAGE = "https://www.istat.it/notizia/basi-territoriali-e-variabili-censuarie/"
LICENSE = "CC BY 3.0 IT"
ANNO = 2021
ETL_VERSION = "0.1.0"

# URL pattern shapefile regionali (R01-R20). R04 = Trentino-Alto Adige incluso.
URL_BT_REGION = (
    "https://www.istat.it/storage/cartografia/basi_territoriali/2021/R{:02d}_21.zip"
)

# Variabili censuarie sezioni: 1 ZIP nazionale con 20 XLSX regionali + TRACCIATO
URL_VARS = (
    "https://esploradati.istat.it/databrowser/DWL/PERMPOP/SUBCOM/"
    "Dati_regionali_2021.zip"
)

# Aree subcomunali (municipi/circoscrizioni/quartieri) - solo ~43 capoluoghi
URL_ASC = "https://www.istat.it/wp-content/uploads/2025/04/ASC_21.zip"

UA = "CruscottoItalia-ETL/1.0 (+https://cruscotto-italia.dati.gov.it)"

# ═════════════════════════════════════════════════════════════════════════
# Lista 122 variabili numeriche estratte per ogni sezione (dal TRACCIATO ISTAT)
# ═════════════════════════════════════════════════════════════════════════

# Popolazione totale e per sesso (3)
VARS_POP_BASE = ["P1", "P2", "P3"]

# Popolazione totale per fascia eta 5 anni (16): P14-P29
VARS_POP_ETA_TOT = [f"P{i}" for i in range(14, 30)]

# Popolazione maschi per fascia eta 5 anni (16): P30-P45
VARS_POP_ETA_M = [f"P{i}" for i in range(30, 46)]

# Popolazione femmine per fascia eta 5 anni (16): P67-P82
VARS_POP_ETA_F = [f"P{i}" for i in range(67, 83)]

# Titolo di studio (18): P83-P100 - totale/M/F x (totale, nessuno, elementare,
# media, diploma, terziario)
VARS_TITOLO = [f"P{i}" for i in range(83, 101)]

# Occupati 15-64 (3): P101-P103 totale/M/F
VARS_OCCUPATI = [f"P{i}" for i in range(101, 104)]

# Italiani (12): IT1-IT12 fasce eta + sesso + occupati
VARS_ITALIANI = [f"IT{i}" for i in range(1, 13)]

# Stranieri (24): ST1, ST2, ST2_B + ST3-ST5 + ST16-ST33
VARS_STRANIERI = (
    ["ST1", "ST2", "ST2_B"]
    + [f"ST{i}" for i in range(3, 6)]
    + [f"ST{i}" for i in range(16, 34)]
)

# Famiglie per numero componenti (7): PF1, PF3-PF8
VARS_FAMIGLIE = ["PF1"] + [f"PF{i}" for i in range(3, 9)]

# Abitazioni ed edifici (4)
VARS_ABITAZIONI = ["A2", "A3", "A8", "E3"]

# Unione totale: 110 codici numerici (i 12 anagrafici come PROCOM, REGIONE
# sono trattati separatamente come metadata della sezione, non in vars{}).
VARS_NUMERIC = (
    VARS_POP_BASE
    + VARS_POP_ETA_TOT
    + VARS_POP_ETA_M
    + VARS_POP_ETA_F
    + VARS_TITOLO
    + VARS_OCCUPATI
    + VARS_ITALIANI
    + VARS_STRANIERI
    + VARS_FAMIGLIE
    + VARS_ABITAZIONI
)
# Sanity check: deve essere 3 + 16 + 16 + 16 + 18 + 3 + 12 + 24 + 7 + 4 = 119
# (le "122 variabili" del titolo includono anche 3 voci anagrafica
#  contestuali: CODREG/CODPRO/CODCOM, gestite come metadata).
assert len(VARS_NUMERIC) == 119, f"Expected 119 vars, got {len(VARS_NUMERIC)}"

# ═════════════════════════════════════════════════════════════════════════
# Mapping dello shapefile TIPO_LOC (campo numerico) -> label leggibile
# ═════════════════════════════════════════════════════════════════════════

TIPO_LOC_MAP = {
    1: "centro_abitato",
    2: "nucleo_abitato",
    3: "case_sparse",
    4: "localita_produttiva",
}

CACHE_DIR = Path("/tmp/cruscotto_censimento")
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# Soglie minime per validare ZIP "non corrotti" (in byte):
# BT regionali piu' piccoli sono ~120KB (es. Molise small), quindi 50KB e' safe
# Variabili Dati_regionali_2021.zip e' ~250MB
MIN_ZIP_SIZE_BT = 50_000
MIN_ZIP_SIZE_VARS = 100_000_000  # 100 MB
MIN_ZIP_SIZE_ASC = 100_000

import time
import urllib.request
import urllib.error

import openpyxl
import shapefile  # pyshp
from pyproj import Transformer


# ═════════════════════════════════════════════════════════════════════════
# FASE 1 — Download fonti (con cache locale per riusabilita')
# ═════════════════════════════════════════════════════════════════════════

def _http_get(url: str, dest: Path, min_size: int, timeout: int = 600) -> Path:
    """Download HTTP con User-Agent + validazione magic bytes ZIP + size minima.

    Cache idempotente: se 'dest' esiste e supera min_size, lo riusa senza
    ri-scaricare. Per forzare il refresh, eliminare il file prima.
    """
    if dest.exists() and dest.stat().st_size >= min_size:
        log.info("censimento_zip_cached",
                 path=str(dest),
                 size_mb=round(dest.stat().st_size / 1024 / 1024, 1))
        return dest

    log.info("censimento_zip_download_start", url=url, dest=str(dest))
    t0 = time.time()
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = resp.read()
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"HTTP {e.code} su {url}: {e.reason}")
    except urllib.error.URLError as e:
        raise RuntimeError(f"URL error su {url}: {e.reason}")
    elapsed = time.time() - t0

    if len(data) < min_size:
        snippet = data[:200].decode("utf-8", errors="ignore")
        raise RuntimeError(
            f"ZIP troppo piccolo ({len(data)} bytes < {min_size}), "
            f"possibile errore HTTP. Snippet: {snippet}"
        )
    if data[:2] != b"PK":
        raise RuntimeError(
            f"Magic bytes non ZIP su {url}: {data[:4].hex()}"
        )

    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(data)
    log.info("censimento_zip_downloaded",
             url=url,
             size_mb=round(len(data) / 1024 / 1024, 1),
             elapsed_s=round(elapsed, 1))
    return dest


def download_dati_regionali(force: bool = False) -> Path:
    """Scarica Dati_regionali_2021.zip (~250MB) contenente 20 XLSX regionali
    di variabili censuarie + 1 XLSX TRACCIATO_2021 con il dizionario.

    Cache: /tmp/cruscotto_censimento/Dati_regionali_2021.zip
    """
    out = CACHE_DIR / "Dati_regionali_2021.zip"
    if force and out.exists():
        out.unlink()
    return _http_get(URL_VARS, out, MIN_ZIP_SIZE_VARS, timeout=900)


def download_bt_region(region: int, force: bool = False) -> Path:
    """Scarica R<NN>_21.zip (shapefile sezioni di censimento) per la regione
    indicata (1-20). R04 = Trentino-Alto Adige.

    Cache: /tmp/cruscotto_censimento/R<NN>_21.zip
    """
    if not 1 <= region <= 20:
        raise ValueError(f"region deve essere 1-20, ricevuto {region}")
    url = URL_BT_REGION.format(region)
    out = CACHE_DIR / f"R{region:02d}_21.zip"
    if force and out.exists():
        out.unlink()
    return _http_get(url, out, MIN_ZIP_SIZE_BT, timeout=300)


def download_asc(force: bool = False) -> Path:
    """Scarica ASC_21.zip (aree subcomunali ~43 capoluoghi, overlay opzionale).

    Cache: /tmp/cruscotto_censimento/ASC_21.zip
    """
    out = CACHE_DIR / "ASC_21.zip"
    if force and out.exists():
        out.unlink()
    return _http_get(URL_ASC, out, MIN_ZIP_SIZE_ASC, timeout=180)


# ═════════════════════════════════════════════════════════════════════════
# FASE 2 — Parse XLSX variabili censuarie (1 file per regione)
# ═════════════════════════════════════════════════════════════════════════

# Codici di colonna anagrafica (NON in VARS_NUMERIC, usati come metadata sezione)
XLSX_META_COLS = ["CODREG", "REGIONE", "CODPRO", "PROVINCIA",
                  "CODCOM", "COMUNE", "PROCOM", "SEZ21_ID",
                  "COM_ASC1", "COM_ASC2", "COM_ASC3"]


def _to_int(val) -> int:
    """Cast robusto a int per celle XLSX: None -> 0, float/str -> int."""
    if val is None:
        return 0
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    s = str(val).strip()
    if not s or s.lower() in ("na", "n.a.", "n/a", "-"):
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def parse_vars_xlsx_from_zip(zip_path: Path, region: int) -> dict[int, dict]:
    """Estrae il file R<NN>_indicatori_2021_sezioni.xlsx dal ZIP nazionale
    e ritorna {SEZ21_ID: {"procom": int, "vars": {P1:..., P2:..., ...}}}.

    Strategia:
    1. Apre il ZIP outer e estrae solo l'XLSX della regione richiesta
       (gli altri 19 file regionali restano non-letti, risparmio memoria)
    2. openpyxl read_only=True data_only=True (no formule, streaming row)
    3. Mappa header -> indice colonna (i nomi sono fissi dal TRACCIATO)
    4. Per ogni riga emette il record con int-cast difensivo

    Note ZIP64: il file nazionale Dati_regionali_2021.zip e' ZIP64 (>4GB
    capability). openpyxl gestisce correttamente i sub-XLSX via zipfile.
    """
    if not 1 <= region <= 20:
        raise ValueError(f"region deve essere 1-20, ricevuto {region}")

    xlsx_name = f"R{region:02d}_indicatori_2021_sezioni.xlsx"
    log.info("censimento_vars_parse_start",
             zip_path=str(zip_path), xlsx=xlsx_name, region=region)
    t0 = time.time()

    # Estrai SOLO il file della regione richiesta in CACHE_DIR per riuso
    xlsx_out = CACHE_DIR / xlsx_name
    if not xlsx_out.exists():
        with zipfile.ZipFile(zip_path, "r") as zf:
            try:
                with zf.open(xlsx_name) as src:
                    xlsx_out.write_bytes(src.read())
            except KeyError:
                raise RuntimeError(
                    f"File {xlsx_name} non trovato nel ZIP. "
                    f"Files disponibili: {zf.namelist()[:5]}"
                )

    wb = openpyxl.load_workbook(xlsx_out, read_only=True, data_only=True)
    ws = wb.active
    log.info("censimento_vars_xlsx_opened",
             region=region, sheet=ws.title, max_row=ws.max_row)

    # Costruisci mapping header -> indice colonna dalla prima riga
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    # Verifica che i campi essenziali esistano
    missing_meta = [c for c in ["PROCOM", "SEZ21_ID"] if c not in col_idx]
    if missing_meta:
        raise RuntimeError(
            f"Colonne meta mancanti nell'XLSX region {region}: {missing_meta}. "
            f"Header trovato (primi 10): {list(col_idx.keys())[:10]}"
        )
    # Conta quante VARS_NUMERIC sono presenti vs mancanti (info, non blocca)
    vars_present = [v for v in VARS_NUMERIC if v in col_idx]
    vars_missing = [v for v in VARS_NUMERIC if v not in col_idx]
    if vars_missing:
        log.warning("censimento_vars_some_missing",
                    region=region, missing=vars_missing[:5],
                    n_missing=len(vars_missing),
                    n_present=len(vars_present))

    out: dict[int, dict] = {}
    n_skipped_no_sez = 0
    for row in rows:
        if row is None:
            continue
        sez_id_raw = row[col_idx["SEZ21_ID"]] if "SEZ21_ID" in col_idx else None
        if sez_id_raw is None:
            n_skipped_no_sez += 1
            continue
        sez_id = _to_int(sez_id_raw)
        if sez_id == 0:
            n_skipped_no_sez += 1
            continue
        procom = _to_int(row[col_idx["PROCOM"]])
        vars_dict = {v: _to_int(row[col_idx[v]]) for v in vars_present}
        out[sez_id] = {"procom": procom, "vars": vars_dict}

    elapsed = time.time() - t0
    log.info("censimento_vars_parse_done",
             region=region,
             sezioni=len(out),
             skipped_no_sez=n_skipped_no_sez,
             vars_present=len(vars_present),
             elapsed_s=round(elapsed, 1))
    return out


def parse_tracciato_from_zip(zip_path: Path) -> dict[str, str]:
    """Estrae il TRACCIATO ISTAT dal ZIP variabili: dict {codice: descrizione}.

    Utile per: generare la tabella collapsibile frontend con la descrizione
    leggibile di ogni codice (es. 'P1' -> 'Popolazione residente - totale').
    """
    tracciato_name = "TRACCIATO FILE REGIONALI.xlsx"
    log.info("censimento_tracciato_parse_start", zip_path=str(zip_path))

    tracciato_out = CACHE_DIR / "TRACCIATO.xlsx"
    if not tracciato_out.exists():
        with zipfile.ZipFile(zip_path, "r") as zf:
            try:
                with zf.open(tracciato_name) as src:
                    tracciato_out.write_bytes(src.read())
            except KeyError:
                log.warning("censimento_tracciato_not_in_zip",
                            zip_path=str(zip_path),
                            available=zf.namelist()[:5])
                return {}

    wb = openpyxl.load_workbook(tracciato_out, read_only=True, data_only=True)
    ws = wb.active
    out: dict[str, str] = {}
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r == 0:
            continue  # header NOME_CAMPO, DEFINIZIONE
        if not row or row[0] is None:
            continue
        code = str(row[0]).strip()
        desc = str(row[1]).strip() if row[1] is not None else ""
        if code:
            out[code] = desc

    log.info("censimento_tracciato_parsed", n_codes=len(out))
    return out


# ═════════════════════════════════════════════════════════════════════════
# FASE 3 — Parse shapefile sezioni di censimento (1 file per regione)
# ═════════════════════════════════════════════════════════════════════════

# Sistema riferimento nativo ISTAT BT2021:
# WGS84 UTM Zona 32N = EPSG:32632 (proiezione metrica, x/y in metri)
# Output target: EPSG:4326 (lon/lat decimali, standard GeoJSON RFC 7946)
SHP_CRS_SRC = "EPSG:32632"
SHP_CRS_DST = "EPSG:4326"

# Tolleranza arrotondamento coordinate riprorettate: 6 decimali = ~11cm
# Risparmia ~30% sul filesize GeoJSON rispetto a 8+ decimali.
COORD_DECIMALS = 6

# Campi DBF attesi nello shapefile R<NN>_21_WGS84.dbf (validati su R02 Valle d'Aosta).
# Se ISTAT cambia layout, l'ETL fallira' con KeyError nel parser (failure loud).
SHP_DBF_FIELDS_REQUIRED = ["PRO_COM", "SEZ21", "SEZ21_ID", "TIPO_LOC"]


def _extract_shp_region(zip_path: Path, region: int) -> Path:
    """Estrae i 4 file shapefile (.shp .dbf .shx .prj) della regione dal ZIP
    BT regionale in una sottodir cache. Ritorna il path al file .shp principale.

    Layout atteso del ZIP regionale (verificato su R02_21.zip):
      SHP/R<NN>_21_WGS84.shp
      SHP/R<NN>_21_WGS84.dbf
      SHP/R<NN>_21_WGS84.shx
      SHP/R<NN>_21_WGS84.prj
      TAB/...  (CSV/XLSX duplicati: ignorati, usiamo Dati_regionali_2021.zip)
    """
    shp_dir = CACHE_DIR / f"R{region:02d}_21_shp"
    shp_main = shp_dir / "SHP" / f"R{region:02d}_21_WGS84.shp"
    if shp_main.exists():
        return shp_main

    shp_dir.mkdir(parents=True, exist_ok=True)
    log.info("censimento_shp_extract_start", region=region, zip=str(zip_path))
    with zipfile.ZipFile(zip_path, "r") as zf:
        # Estrai solo i 4 file SHP/ (no TAB/ duplicati)
        prefix = "SHP/"
        n_extracted = 0
        for name in zf.namelist():
            if name.startswith(prefix) and not name.endswith("/"):
                zf.extract(name, shp_dir)
                n_extracted += 1
        log.info("censimento_shp_extracted", region=region, files=n_extracted)

    if not shp_main.exists():
        raise RuntimeError(
            f"File shapefile principale non trovato dopo estrazione: {shp_main}"
        )
    return shp_main


def parse_shapefile_region(zip_path: Path, region: int) -> list[dict]:
    """Estrae e parsa lo shapefile sezioni della regione.

    Ritorna lista di dict per ogni sezione:
      {
        "sez_id": int (SEZ21_ID, chiave globale univoca),
        "procom": str (codice ISTAT 6 cifre PRO_COM zero-padded),
        "sez": int (SEZ21, progressivo locale al comune),
        "tipo_loc": str (label da TIPO_LOC_MAP, default "altro"),
        "asc1": int, "asc2": int, "asc3": int (codici aree subcomunali),
        "area_mq": float (SHAPE_Area dal DBF),
        "geom": dict GeoJSON Polygon EPSG:4326 (lon/lat WGS84)
      }

    Le geometrie sono riproiettate da UTM 32N a WGS84 lon/lat e arrotondate
    a 6 decimali (~11cm). I poligoni con buchi (MultiPolygon-like via 'parts')
    sono gestiti come rings multiple del Polygon GeoJSON.
    """
    shp_path = _extract_shp_region(zip_path, region)
    log.info("censimento_shp_parse_start", region=region, shp=str(shp_path))
    t0 = time.time()

    sf = shapefile.Reader(str(shp_path), encoding="utf-8")
    transformer = Transformer.from_crs(SHP_CRS_SRC, SHP_CRS_DST, always_xy=True)

    # Verifica colonne DBF essenziali
    field_names = [f[0] for f in sf.fields[1:]]  # skip DeletionFlag
    missing = [c for c in SHP_DBF_FIELDS_REQUIRED if c not in field_names]
    if missing:
        raise RuntimeError(
            f"DBF region {region}: campi mancanti {missing}. "
            f"Trovati (primi 10): {field_names[:10]}"
        )

    out: list[dict] = []
    n_skipped_no_geom = 0
    n_skipped_bad_geom = 0

    for shape, rec in zip(sf.shapes(), sf.records()):
        d = rec.as_dict()

        # Skip se shape non e' Polygon (5) o MultiPatch (31).
        # ISTAT BT2021 sono tutti Polygon ma difensivi.
        if shape.shapeType != shapefile.POLYGON:
            n_skipped_bad_geom += 1
            continue
        if not shape.points or len(shape.points) < 4:
            # Polygon valido ha almeno 4 punti (ring chiuso)
            n_skipped_no_geom += 1
            continue

        # Riproietta ring per ring (shape.parts e' la lista di start-index dei ring)
        parts = list(shape.parts) + [len(shape.points)]
        rings = []
        for i in range(len(parts) - 1):
            start, end = parts[i], parts[i + 1]
            ring = []
            for x, y in shape.points[start:end]:
                lon, lat = transformer.transform(x, y)
                ring.append([round(lon, COORD_DECIMALS), round(lat, COORD_DECIMALS)])
            rings.append(ring)

        # ISTAT campo PRO_COM e' int (es. 7001 = Aosta), padded a 6 cifre come istat
        procom_int = int(d.get("PRO_COM") or 0)
        procom_str = f"{procom_int:06d}"

        out.append({
            "sez_id": int(d.get("SEZ21_ID") or 0),
            "procom": procom_str,
            "sez": int(d.get("SEZ21") or 0),
            "tipo_loc": TIPO_LOC_MAP.get(int(d.get("TIPO_LOC") or 0), "altro"),
            "loc_id": int(d.get("LOC21_ID") or 0),
            "asc1": int(d.get("COM_ASC1") or 0),
            "asc2": int(d.get("COM_ASC2") or 0),
            "asc3": int(d.get("COM_ASC3") or 0),
            "area_mq": round(float(d.get("SHAPE_Area") or 0), 1),
            "geom": {"type": "Polygon", "coordinates": rings},
        })

    elapsed = time.time() - t0
    log.info("censimento_shp_parse_done",
             region=region,
             features=len(out),
             skipped_no_geom=n_skipped_no_geom,
             skipped_bad_geom=n_skipped_bad_geom,
             elapsed_s=round(elapsed, 1))
    return out


# ═════════════════════════════════════════════════════════════════════════
# FASE 4 — Aggregati comune-level (KPI + distribuzioni)
# ═════════════════════════════════════════════════════════════════════════
# Vengono calcolati a partire dalle vars{} delle sezioni di ogni comune e
# scritti in data/censimento/<istat>.json (sezione "censimento" del
# dashboard A1). Le geometrie complete vanno in data/censimento_full/.

# Mapping codice fascia -> label e codici sommabili per le distribuzioni.
# Allineato al TRACCIATO ISTAT (vedi docstring del modulo).
ETA_5ANNI_BUCKETS = [
    ("0-4", "P14"), ("5-9", "P15"), ("10-14", "P16"), ("15-19", "P17"),
    ("20-24", "P18"), ("25-29", "P19"), ("30-34", "P20"), ("35-39", "P21"),
    ("40-44", "P22"), ("45-49", "P23"), ("50-54", "P24"), ("55-59", "P25"),
    ("60-64", "P26"), ("65-69", "P27"), ("70-74", "P28"), ("75+", "P29"),
]
ETA_FASCE_AGGR = {
    "0-14": ["P14", "P15", "P16"],
    "15-64": ["P17", "P18", "P19", "P20", "P21", "P22",
              "P23", "P24", "P25", "P26"],
    "65+": ["P27", "P28", "P29"],
}
TITOLO_BUCKETS = [
    ("nessuno", "P86"), ("elementare", "P87"), ("media", "P88"),
    ("diploma", "P89"), ("terziario", "P90"),
]
FAMIGLIE_BUCKETS = [
    ("1", "PF3"), ("2", "PF4"), ("3", "PF5"),
    ("4", "PF6"), ("5", "PF7"), ("6+", "PF8"),
]
STRANIERI_ETA_BUCKETS = [
    ("0-29", "ST3"), ("30-54", "ST4"), ("55+", "ST5"),
]


def aggregate_comune(features_with_vars: list[dict]) -> dict:
    """Costruisce il dict aggregato comune-level (KPI + distribuzioni)
    a partire dalla lista di sezioni (ognuna con 'vars' dict).

    Output: dict serializzabile JSON pronto per essere scritto in
    data/censimento/<istat>.json (sezione 'censimento' del dashboard A1).

    NON include le geometrie (che vanno in censimento_full/).
    """
    def sumv(code: str) -> int:
        """Somma il codice variabile su tutte le sezioni (default 0)."""
        return sum(int(f.get("vars", {}).get(code, 0) or 0)
                   for f in features_with_vars)

    n_sezioni = len(features_with_vars)
    area_kmq = round(
        sum(f.get("area_mq", 0) for f in features_with_vars) / 1_000_000, 3
    )

    kpi_comune = {
        "n_sezioni": n_sezioni,
        "pop_totale": sumv("P1"),
        "pop_maschi": sumv("P2"),
        "pop_femmine": sumv("P3"),
        "famiglie_totali": sumv("PF1"),
        "abitazioni_totali": sumv("A8"),
        "abitazioni_occupate": sumv("A2"),
        "abitazioni_vuote": sumv("A3"),
        "edifici_residenziali": sumv("E3"),
        "stranieri_totali": sumv("ST1"),
        "stranieri_ue": sumv("ST16"),
        "stranieri_extra_ue": sumv("ST19"),
        "occupati_15_64": sumv("P101"),
        "occupati_maschi": sumv("P102"),
        "occupati_femmine": sumv("P103"),
        "area_kmq": area_kmq,
    }

    distribuzioni = {
        "eta_5anni": {label: sumv(code) for label, code in ETA_5ANNI_BUCKETS},
        "eta_per_fascia": {label: sum(sumv(c) for c in codes)
                           for label, codes in ETA_FASCE_AGGR.items()},
        "titolo_studio_9plus": {label: sumv(code) for label, code in TITOLO_BUCKETS},
        "famiglie_componenti": {label: sumv(code) for label, code in FAMIGLIE_BUCKETS},
        "stranieri_eta": {label: sumv(code) for label, code in STRANIERI_ETA_BUCKETS},
    }

    return {
        "_etl_version": ETL_VERSION,
        "_source": SOURCE_LABEL,
        "_source_url": SOURCE_PAGE,
        "_license": LICENSE,
        "_anno_rilevazione": ANNO,
        "_generated_at": datetime.now(timezone.utc).isoformat(),
        "_has_full": True,   # i poligoni completi sono in /data/censimento_full/<istat>.geojson
        "_has_asc": False,   # popolato dall'ETL ASC separato
        "kpi_comune": kpi_comune,
        "distribuzioni_comune": distribuzioni,
    }


# ═════════════════════════════════════════════════════════════════════════
# FASE 5 — Main ETL: loop 20 regioni + scrittura GeoJSON + aggregati
# ═════════════════════════════════════════════════════════════════════════

DEFAULT_FULL_DIR = "/var/www/cruscotto-italia/data/censimento_full"
DEFAULT_AGG_DIR = "/var/www/cruscotto-italia/data/censimento"


def _write_atomic_json(path: Path, payload, *, indent: int | None = None) -> None:
    """Scrive JSON atomic (tmp + rename) per evitare letture parziali se
    nginx serve il file mentre l'ETL e' in corso.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    if indent is None:
        text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    else:
        text = json.dumps(payload, ensure_ascii=False, indent=indent)
    tmp.write_text(text, encoding="utf-8")
    tmp.replace(path)


def run_etl(
    regions: list[int],
    full_outdir: Path,
    agg_outdir: Path,
    keep_cache: bool = True,
) -> dict:
    """Esegue l'ETL completo per le regioni indicate.

    Sequenza:
    1. Verifica/scarica Dati_regionali_2021.zip (vars nazionale).
    2. Per ogni regione:
       a. Download R<NN>_21.zip (shapefile BT).
       b. Parse XLSX variabili regione -> {sez_id: {procom, vars}}.
       c. Parse shapefile -> [{sez_id, procom, geom, ...}].
       d. Merge: arricchisce ogni feature con il dict 'vars'.
    3. Raggruppa tutte le sezioni per procom (codice ISTAT 6 cifre).
    4. Per ogni comune scrive:
       - {full_outdir}/<istat>.geojson : FeatureCollection sezioni + vars
       - {agg_outdir}/<istat>.json     : KPI + distribuzioni aggregate

    Ritorna stats dict per il manifest update.
    """
    full_outdir = Path(full_outdir)
    agg_outdir = Path(agg_outdir)
    full_outdir.mkdir(parents=True, exist_ok=True)
    agg_outdir.mkdir(parents=True, exist_ok=True)

    log.info("censimento_etl_start",
             regions=regions,
             full_outdir=str(full_outdir),
             agg_outdir=str(agg_outdir))
    t_total = time.time()

    # 1. Vars nazionale (download SOLO se manca; ~250MB cache idempotente)
    vars_zip = download_dati_regionali()

    # 2. Loop regioni, costruisco lista globale sezioni con vars+geom
    all_sezioni: list[dict] = []
    per_region_stats: list[dict] = []

    for region in regions:
        t_reg = time.time()
        log.info("censimento_region_start", region=region)

        bt_zip = download_bt_region(region)
        vars_by_sezid = parse_vars_xlsx_from_zip(vars_zip, region)
        geo_features = parse_shapefile_region(bt_zip, region)

        # Merge sez_id per sez_id
        n_merged, n_no_vars = 0, 0
        for feat in geo_features:
            rec = vars_by_sezid.get(feat["sez_id"])
            if rec is None:
                n_no_vars += 1
                feat["vars"] = {}
            else:
                feat["vars"] = rec["vars"]
                n_merged += 1
            all_sezioni.append(feat)

        log.info("censimento_region_merged",
                 region=region,
                 n_features=len(geo_features),
                 n_merged=n_merged,
                 n_no_vars=n_no_vars,
                 elapsed_s=round(time.time() - t_reg, 1))
        per_region_stats.append({
            "region": region,
            "n_features": len(geo_features),
            "n_merged": n_merged,
            "n_no_vars": n_no_vars,
        })

    # 3. Raggruppa per procom
    by_procom: dict[str, list[dict]] = defaultdict(list)
    for sez in all_sezioni:
        by_procom[sez["procom"]].append(sez)
    log.info("censimento_grouped_by_procom",
             total_sezioni=len(all_sezioni),
             n_comuni=len(by_procom))

    # 4. Validazione anagrafica (best-effort, non blocca)
    try:
        from etl.lib import local_lookup as _ll
        bundle = _ll.load_comuni_bundle()
        valid_istat = set(bundle.keys()) if bundle else None
    except Exception as e:
        log.warning("censimento_anagrafica_load_failed", error=str(e))
        valid_istat = None
    n_istat_not_in_bundle = 0
    if valid_istat is not None:
        missing_in_bundle = [p for p in by_procom if p not in valid_istat]
        n_istat_not_in_bundle = len(missing_in_bundle)
        if missing_in_bundle:
            log.warning("censimento_istat_not_in_bundle",
                        n=n_istat_not_in_bundle,
                        sample=missing_in_bundle[:5])

    # 5. Scrittura per comune (GeoJSON full + aggregati)
    t_write = time.time()
    n_written_full, n_written_agg = 0, 0
    sample_files = []
    for procom, sezioni in by_procom.items():
        # 5a. GeoJSON FULL (geometrie + vars)
        features_geojson = []
        for s in sezioni:
            features_geojson.append({
                "type": "Feature",
                "properties": {
                    "id": s["sez_id"],
                    "sez": s["sez"],
                    "tipo_loc": s["tipo_loc"],
                    "loc_id": s["loc_id"],
                    "asc1": s["asc1"],
                    "asc2": s["asc2"],
                    "asc3": s["asc3"],
                    "area_mq": s["area_mq"],
                    "vars": s.get("vars", {}),
                },
                "geometry": s["geom"],
            })
        fc = {
            "type": "FeatureCollection",
            "_source": SOURCE_LABEL,
            "_istat": procom,
            "_n_sezioni": len(sezioni),
            "_generated_at": datetime.now(timezone.utc).isoformat(),
            "features": features_geojson,
        }
        full_path = full_outdir / f"{procom}.geojson"
        _write_atomic_json(full_path, fc)
        n_written_full += 1

        # 5b. Aggregati comune-level (per dashboard A1)
        agg = aggregate_comune(sezioni)
        agg_path = agg_outdir / f"{procom}.json"
        _write_atomic_json(agg_path, agg)
        n_written_agg += 1

        # Tiene 3 sample per log diagnostico finale
        if len(sample_files) < 3:
            sample_files.append({
                "istat": procom,
                "n_sezioni": len(sezioni),
                "full_size_kb": round(full_path.stat().st_size / 1024, 1),
                "agg_size_kb": round(agg_path.stat().st_size / 1024, 1),
            })

    log.info("censimento_write_done",
             n_full=n_written_full,
             n_agg=n_written_agg,
             elapsed_s=round(time.time() - t_write, 1),
             samples=sample_files)

    elapsed_total = round(time.time() - t_total, 1)
    stats = {
        "regions": regions,
        "n_comuni": len(by_procom),
        "n_sezioni_totali": len(all_sezioni),
        "n_istat_not_in_bundle": n_istat_not_in_bundle,
        "per_region": per_region_stats,
        "elapsed_s": elapsed_total,
    }
    log.info("censimento_etl_done", **{k: v for k, v in stats.items()
                                       if k != "per_region"})
    return stats


# ═════════════════════════════════════════════════════════════════════════
# SMOKE TEST — esegui con: python -m etl.sources.censimento --smoke-test
# ═════════════════════════════════════════════════════════════════════════
# Scarica solo Valle d'Aosta (regione 2, ZIP piu' piccolo ~4MB BT + ~1MB
# XLSX) e Cogne come comune campione per validare la pipeline su dato vero
# senza scaricare i 250MB del file nazionale variabili.

def _smoke_test() -> int:
    """Test end-to-end veloce solo su Valle d'Aosta + comune Cogne.

    Pre-condizione: ZIP nazionale Dati_regionali_2021.zip gia' presente in
    CACHE_DIR (e' 250MB, non lo scarichiamo nello smoke). Per eseguirlo
    la prima volta, lanciare prima con --download-vars.

    Output atteso:
    - Region 2 (VdA): 4485 sezioni totali
    - Comune Cogne (PRO_COM=7021, istat=007021): ~12 sezioni
    - Sample sezione: 1 record completo con vars + geom
    """
    region = 2  # Valle d'Aosta
    cogne_istat = "007021"  # PRO_COM=7021 padded

    print("=" * 70)
    print("SMOKE TEST CENSIMENTO ISTAT - Valle d'Aosta / Cogne")
    print("=" * 70)

    # 1. Verifica/scarica BT regione 2
    print(f"\n[1/4] Download BT regione R02_21.zip...")
    bt_zip = download_bt_region(region)
    print(f"      OK: {bt_zip} ({bt_zip.stat().st_size / 1024 / 1024:.1f} MB)")

    # 2. Verifica file vars nazionale (pre-condizione)
    vars_zip = CACHE_DIR / "Dati_regionali_2021.zip"
    if not vars_zip.exists():
        print(f"\n  ATTENZIONE: {vars_zip} non trovato.")
        print(f"  Per popolare la cache, lancia prima:")
        print(f"    python3 -m etl.sources.censimento --download-vars")
        print(f"  (250 MB download, ~5-10 min). Smoke test interrotto.")
        return 1
    print(f"\n[2/4] File variabili nazionale: cache hit OK "
          f"({vars_zip.stat().st_size / 1024 / 1024:.1f} MB)")

    # 3. Parse XLSX variabili regione 2
    print(f"\n[3/4] Parse XLSX variabili R02...")
    vars_by_sezid = parse_vars_xlsx_from_zip(vars_zip, region)
    print(f"      OK: {len(vars_by_sezid)} sezioni con variabili")
    # Estraggo le sezioni di Cogne (PRO_COM == 7021)
    cogne_sezid = {sid for sid, rec in vars_by_sezid.items()
                   if rec["procom"] == 7021}
    print(f"      Sezioni Cogne (PRO_COM=7021): {len(cogne_sezid)}")

    # 4. Parse shapefile regione 2
    print(f"\n[4/4] Parse shapefile R02...")
    features = parse_shapefile_region(bt_zip, region)
    print(f"      OK: {len(features)} sezioni totali in VdA")
    cogne_features = [f for f in features if f["procom"] == cogne_istat]
    print(f"      Sezioni Cogne in SHP: {len(cogne_features)}")

    # 5. Cross-check sez_id geometrie vs vars
    geom_sezids = {f["sez_id"] for f in cogne_features}
    only_in_geom = geom_sezids - cogne_sezid
    only_in_vars = cogne_sezid - geom_sezids
    print(f"\n[CROSS] sez_id Cogne in geometrie: {len(geom_sezids)}, "
          f"in vars: {len(cogne_sezid)}")
    if only_in_geom:
        print(f"        Solo in geometrie (no vars): {len(only_in_geom)} -> "
              f"{sorted(only_in_geom)[:5]}")
    if only_in_vars:
        print(f"        Solo in vars (no geometria): {len(only_in_vars)} -> "
              f"{sorted(only_in_vars)[:5]}")

    # 6. Sample sezione completa (la prima di Cogne con vars+geom)
    if cogne_features:
        sample = cogne_features[0]
        sid = sample["sez_id"]
        vars_rec = vars_by_sezid.get(sid, {}).get("vars", {})
        print(f"\n[SAMPLE] Prima sezione Cogne:")
        print(f"  sez_id:    {sid}")
        print(f"  sez:       {sample['sez']}")
        print(f"  procom:    {sample['procom']}")
        print(f"  tipo_loc:  {sample['tipo_loc']}")
        print(f"  area_mq:   {sample['area_mq']}")
        print(f"  geom rings: {len(sample['geom']['coordinates'])}, "
              f"points ring0: {len(sample['geom']['coordinates'][0])}")
        print(f"  primo punto: {sample['geom']['coordinates'][0][0]}  (lon, lat)")
        print(f"  vars (first 5): "
              f"{dict(list(vars_rec.items())[:5]) if vars_rec else 'NESSUNA'}")
        print(f"  P1 popolazione: {vars_rec.get('P1', 'N/A')}")
        print(f"  P101 occupati 15-64: {vars_rec.get('P101', 'N/A')}")
        print(f"  ST1 stranieri: {vars_rec.get('ST1', 'N/A')}")
        print(f"  PF1 famiglie: {vars_rec.get('PF1', 'N/A')}")
        print(f"  A8 abitazioni tot: {vars_rec.get('A8', 'N/A')}")
    else:
        print(f"\n[ERROR] Nessuna sezione Cogne nelle geometrie!")
        return 2

    print(f"\n{'=' * 70}\nSMOKE TEST OK\n{'=' * 70}")
    return 0


def _download_vars_only() -> int:
    """Scarica solo Dati_regionali_2021.zip (250MB) nella cache.
    Necessario una sola volta prima dello smoke-test.
    """
    print("Download Dati_regionali_2021.zip (~250 MB, puo' richiedere 5-10 min)...")
    out = download_dati_regionali()
    print(f"OK: {out} ({out.stat().st_size / 1024 / 1024:.1f} MB)")
    return 0


if __name__ == "__main__":
    import sys
    parser = argparse.ArgumentParser(
        description="ETL ISTAT Basi Territoriali + Variabili censuarie 2021"
    )
    parser.add_argument(
        "--target", choices=["local"], default="local",
        help="Target output (solo local supportato post local-first refactor)"
    )
    parser.add_argument(
        "--regions", default="all",
        help="Lista regioni 1-20 separate da virgola, o 'all' (default)"
    )
    parser.add_argument(
        "--full-outdir", default=DEFAULT_FULL_DIR,
        help=f"Directory GeoJSON full (default: {DEFAULT_FULL_DIR})"
    )
    parser.add_argument(
        "--agg-outdir", default=DEFAULT_AGG_DIR,
        help=f"Directory aggregati comune-level (default: {DEFAULT_AGG_DIR})"
    )
    parser.add_argument(
        "--download-vars", action="store_true",
        help="Scarica solo Dati_regionali_2021.zip nella cache, poi esce"
    )
    parser.add_argument(
        "--smoke-test", action="store_true",
        help="Smoke test su Valle d'Aosta + Cogne, poi esce"
    )
    args = parser.parse_args()

    if args.download_vars:
        sys.exit(_download_vars_only())
    if args.smoke_test:
        sys.exit(_smoke_test())

    # Parse regioni
    if args.regions == "all":
        regions = list(range(1, 21))
    else:
        try:
            regions = [int(r.strip()) for r in args.regions.split(",")]
        except ValueError as e:
            print(f"ERROR: --regions invalido: {e}", file=sys.stderr)
            sys.exit(2)
        for r in regions:
            if not 1 <= r <= 20:
                print(f"ERROR: regione {r} fuori range 1-20", file=sys.stderr)
                sys.exit(2)

    stats = run_etl(
        regions=regions,
        full_outdir=Path(args.full_outdir),
        agg_outdir=Path(args.agg_outdir),
    )

    # Manifest update
    try:
        files_entry = [
            {"key": f"censimento/{stats['n_comuni']}_comuni",
             "n_comuni": stats["n_comuni"],
             "n_sezioni_totali": stats["n_sezioni_totali"],
             "regioni_processate": stats["regions"]},
            {"key": f"censimento_full/{stats['n_comuni']}_geojson"},
        ]
        manifest.update_source("censimento", files_entry, status="ok")
    except Exception as e:
        log.error("censimento_manifest_update_failed", error=str(e))

    sys.exit(0)
